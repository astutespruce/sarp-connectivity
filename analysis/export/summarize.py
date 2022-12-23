from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, NamedStyle
from openpyxl.utils import get_column_letter

from api.constants import DOMAINS
from analysis.constants import SARP_STATES

primary_col_style = NamedStyle(
    name="PrimaryColumnStyle", alignment=Alignment(horizontal="left", wrap_text=True)
)


current_version = "Dec2022"

data_dir = Path("data/barriers/master")
out_dir = Path("data/versions")

had_manual_review = [4, 5, 8, 10, 11, 13, 14, 15]

common_status_fields = [
    "snapped",
    "excluded",
    "dropped",
    "duplicate",
    "unranked",
    # "removed",
    "loop",
    "ProtectedLand",
]
barrier_status_fields = {
    "dams": common_status_fields
    + [
        "is_estimated",
    ],
    "small_barriers": common_status_fields.copy(),
}


common_summary_fields = [
    "Recon",
    "ManualReview",
    "BarrierSeverity",
    "Condition",
    "BarrierOwnerType",
    "OwnerType",
    "removed",
    "log",
    "snap_log",
]
barrier_summary_fields = {
    "dams": common_summary_fields,
    "small_barriers": common_summary_fields + ["PotentialProject", "CrossingType"],
}


for barrier_type in ["dams", "small_barriers"]:
    print(f"Summarizing {barrier_type}...")
    type_label = "dams" if barrier_type == "dams" else "road-related barriers"

    status_fields = barrier_status_fields[barrier_type]
    summary_fields = barrier_summary_fields[barrier_type]

    extra_fields = ["SARP_Score"] if barrier_type == "small_barriers" else []

    df = pd.read_feather(
        data_dir / f"{barrier_type}.feather",
        columns=["id"]
        + summary_fields
        + status_fields
        + extra_fields
        + ["State", "HUC2"],
    )
    api_df = pd.read_feather(
        f"data/api/{barrier_type}.feather",
        columns=[
            "id",
            "HasNetwork",
            "Ranked",
        ],
    ).set_index("id")
    df.loc[df.State == "", "State"] = "<outside region states>"

    df = df.join(api_df, on="id")
    df["not_dropped_or_duplicate_or_removed"] = (
        ~(df.dropped | df.duplicate | df.removed)
    ).astype("uint8")
    df["analyzed"] = (df.snapped & ~(df.duplicate | df.dropped | df.excluded)).astype(
        "uint8"
    )
    df["manually_reviewed"] = df.ManualReview.isin(had_manual_review).astype("uint8")
    df["reconned"] = (df.Recon > 0).astype("uint8")
    df["removed_not_dropped_or_duplicate"] = df.removed & (~(df.dropped | df.duplicate))

    if barrier_type == "dams":
        # per guidance from Kat: confirmed unless within SARP states and not reconned or is error
        df["is_confirmed_estimated"] = df.is_estimated & ~(
            df.State.isin(SARP_STATES)
            & (df.Recon.isin([0, 5]) | df.ManualReview.isin([6, 11, 14]))
        ).astype("uint8")
    else:
        df["proposed_project"] = df.PotentialProject.isin(["Proposed Project"]).astype(
            "uint8"
        )
        df["scored"] = df.SARP_Score != -1

    for col in status_fields + [
        "HasNetwork",
        "Ranked",
        "ProtectedLand",
        "OwnerType",
        "BarrierOwnerType",
    ]:
        df[col] = df[col].fillna(0).astype("uint8")

    states = sorted(x for x in df.State.unique() if x)
    huc2 = sorted(x for x in df.HUC2.unique() if x)

    with pd.ExcelWriter(
        out_dir / f"{barrier_type}_{current_version}_summary.xlsx"
    ) as xlsx, open(
        out_dir / f"{barrier_type}_{current_version}_summary.md", "w"
    ) as md:

        ### Totals
        data = {
            "total": len(df),
            "not_dropped_or_duplicate_or_removed": df.not_dropped_or_duplicate_or_removed.sum(),
            "reconned": df.reconned.sum(),
            "manually_reviewed": df.manually_reviewed.sum(),
        }

        if barrier_type == "dams":
            data["estimated"] = df.is_estimated.sum()
            data[
                "confirmed_estimated (not in SARP states or in SARP state & not error)"
            ] = df.is_confirmed_estimated.sum()
        else:
            data["scored (SARP_Score>=0)"] = df.scored.sum()
            data["proposed_project"] = df.proposed_project.sum()
            summary_fields += ["proposed_project", "scored"]

        data.update(
            {
                "dropped (errors, etc)": df.dropped.sum(),
                "excluded (do not cut network)": df.excluded.sum(),
                "duplicate": df.duplicate.sum(),
                "snapped": df.snapped.sum(),
                "on_loop": df.loop.sum(),
                "removed": df.removed.sum(),
                "removed_not_dropped_or_duplicate": df.removed_not_dropped_or_duplicate.sum(),
                "analyzed (snapped & not excluded/dropped/duplicate)": df.analyzed.sum(),
                "has_networkresults": df.HasNetwork.sum(),
                "unranked (invasive)": (df.unranked == 1).sum(),
                "ranked": df.Ranked.sum(),
                "on_protectedland": df.ProtectedLand.sum(),
            }
        )

        total = pd.DataFrame([data]).melt()
        total.columns = ["", "count"]
        sheet_name = f"All {type_label}"
        total.to_excel(xlsx, sheet_name=sheet_name, index=False)
        ws = xlsx.sheets[sheet_name]
        ws.column_dimensions["A"].width = 32

        md.write(f"## All {type_label}\n")
        md.write(total.to_markdown(floatfmt=",.0f", index=False))
        md.write("\n\n\n\n")

        ### Totals by state
        agg = {
            "id": "count",
            "not_dropped_or_duplicate_or_removed": "sum",
            "reconned": "sum",
            "manually_reviewed": "sum",
        }

        if barrier_type == "dams":
            agg["is_estimated"] = "sum"
            agg["is_confirmed_estimated"] = "sum"
        else:
            agg["scored"] = "sum"
            agg["proposed_project"] = "sum"

        agg.update(
            {
                "dropped": "sum",
                "excluded": "sum",
                "duplicate": "sum",
                "snapped": "sum",
                "loop": "sum",
                "removed": "sum",
                "removed_not_dropped_or_duplicate": "sum",
                "analyzed": "sum",
                "HasNetwork": "sum",
                "unranked": "sum",
                "Ranked": "sum",
                "ProtectedLand": "sum",
            }
        )

        rename_cols = {
            "id": "total",
            "ProtectedLand": "protectedland",
            "is_estimated": "estimated",
            "is_confirmed_estimated": "confirmed_estimated",
            "HasNetwork": "has_networkresults",
            "Ranked": "ranked",
            "loop": "on_loop",
        }

        stats = df.groupby("State").agg(agg).rename(columns=rename_cols)
        sheet_name = f"{type_label} by state"
        stats.to_excel(xlsx, sheet_name=sheet_name)
        ws = xlsx.sheets[sheet_name]
        ws.column_dimensions["A"].width = 24
        for idx in range(2, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 16

        for row in list(ws.rows)[1:]:
            row[0].style = primary_col_style

        md.write(f"## {type_label} by state\n")
        md.write(stats.to_markdown(floatfmt=",.0f"))
        md.write("\n\n\n\n")

        ### Totals by HUC2
        stats = df.groupby("HUC2").agg(agg).rename(columns=rename_cols)
        sheet_name = f"{type_label} by HUC2"
        stats.to_excel(xlsx, sheet_name=sheet_name)
        ws = xlsx.sheets[sheet_name]
        ws.column_dimensions["A"].width = 6
        for idx in range(2, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 16

        for row in list(ws.rows)[1:]:
            row[0].style = primary_col_style

        md.write(f"## {type_label} by HUC2\n")
        md.write(stats.to_markdown(floatfmt=",.0f"))
        md.write("\n\n\n\n")

        for col in summary_fields:
            agg = {"id": "count", "not_dropped_or_duplicate_or_removed": "sum"}

            if col != "Recon":
                agg["reconned"] = "sum"

            if col != "ManualReview":
                agg["manually_reviewed"] = "sum"

            if col != "removed":
                agg["removed"] = "sum"
                agg["removed_not_dropped_or_duplicate"] = "sum"

            if barrier_type == "dams":
                agg["is_estimated"] = "sum"
                agg["is_confirmed_estimated"] = "sum"

            else:
                if col != "proposed_project":
                    agg["proposed_project"] = "sum"

                if col != "scored":
                    agg["scored"] = "sum"

            agg.update(
                {
                    "dropped": "sum",
                    "excluded": "sum",
                    "duplicate": "sum",
                    "snapped": "sum",
                    "loop": "sum",
                    "analyzed": "sum",
                    "HasNetwork": "sum",
                    "unranked": "sum",
                    "Ranked": "sum",
                    "ProtectedLand": "sum",
                }
            )

            rename_cols = {
                "id": "total",
                "ProtectedLand": "protectedland",
                "is_estimated": "estimated",
                "is_confirmed_estimated": "confirmed_estimated",
                "HasNetwork": "has_networkresults",
                "Ranked": "ranked",
                "loop": "on_loop",
            }

            stats = df.groupby(col).agg(agg).rename(columns=rename_cols)

            if col in DOMAINS:
                stats.index = (
                    stats.index.astype(str)
                    + ": "
                    + stats.index.map(DOMAINS[col]).fillna("")
                )

            sheet_name = f"{col} {type_label}"
            stats.to_excel(xlsx, sheet_name=sheet_name)
            ws = xlsx.sheets[sheet_name]
            ws.column_dimensions["A"].width = 60
            for idx in range(2, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(idx)].width = 16

            for row in list(ws.rows)[1:]:
                row[0].style = primary_col_style

            md.write(f"## {col} {type_label}\n")
            md.write(stats.to_markdown(floatfmt=",.0f"))
            md.write("\n\n\n\n")

            ### Crosstab by state
            values = df[col].copy()
            if col in DOMAINS:
                values = (
                    df[col].astype(str) + ": " + df[col].map(DOMAINS[col]).fillna("")
                )

            crosstab = pd.crosstab(
                values,
                df.State,
            )

            sheet_name = f"{col} {type_label} state crosstab (total)"
            crosstab.to_excel(xlsx, sheet_name=sheet_name)
            ws = xlsx.sheets[sheet_name]
            ws.column_dimensions["A"].width = 60
            for idx in range(2, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(idx)].width = 16

            for row in list(ws.rows)[1:]:
                row[0].style = primary_col_style

            md.write(f"## {col} {type_label} state crosstab (total)\n")
            md.write(crosstab.to_markdown(floatfmt=",.0f"))
            md.write("\n\n\n\n")

            ### Values by state
            stats = (
                df.groupby(["State", col])
                .agg(agg)
                .rename(columns=rename_cols)
                .reset_index()
            )
            if col in DOMAINS:
                stats[col] = (
                    stats[col].astype(str)
                    + ": "
                    + stats[col].map(DOMAINS[col]).fillna("")
                )

            sheet_name = f"{col} {type_label} by state"
            stats.to_excel(xlsx, sheet_name=sheet_name, index=False)
            ws = xlsx.sheets[sheet_name]
            ws.column_dimensions["A"].width = 24
            ws.column_dimensions["B"].width = 60
            for idx in range(3, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(idx)].width = 16

            for row in list(ws.rows)[1:]:
                row[0].style = primary_col_style
                row[1].style = primary_col_style

            md.write(f"## {col} {type_label} by state\n")
            md.write(stats.to_markdown(floatfmt=",.0f", index=False))
            md.write("\n\n\n\n")

            ### Crosstab by HUC2
            values = df[col].copy()
            if col in DOMAINS:
                values = (
                    df[col].astype(str) + ": " + df[col].map(DOMAINS[col]).fillna("")
                )

            crosstab = pd.crosstab(values, df.HUC2)

            sheet_name = f"{col} {type_label} HUC2 crosstab (total)"
            crosstab.to_excel(xlsx, sheet_name=sheet_name)
            ws = xlsx.sheets[sheet_name]
            ws.column_dimensions["A"].width = 60
            for idx in range(2, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(idx)].width = 16

            for row in list(ws.rows)[1:]:
                row[0].style = primary_col_style

            md.write(f"## {col} {type_label} HUC2 crosstab (total)\n")
            md.write(crosstab.to_markdown(floatfmt=",.0f"))
            md.write("\n\n\n\n")

            ### Values by HUC2
            stats = (
                df.groupby(["HUC2", col])
                .agg(agg)
                .rename(columns=rename_cols)
                .reset_index()
            )
            if col in DOMAINS:
                stats[col] = (
                    stats[col].astype(str)
                    + ": "
                    + stats[col].map(DOMAINS[col]).fillna("")
                )

            sheet_name = f"{col} {type_label} by HUC2"
            stats.to_excel(xlsx, sheet_name=sheet_name, index=False)
            ws = xlsx.sheets[sheet_name]
            ws.column_dimensions["A"].width = 6
            ws.column_dimensions["B"].width = 60
            for idx in range(3, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(idx)].width = 16

            for row in list(ws.rows)[1:]:
                row[0].style = primary_col_style
                row[1].style = primary_col_style

            md.write(f"## {col} {type_label} by HUC2\n")
            md.write(stats.to_markdown(floatfmt=",.0f", index=False))
            md.write("\n\n\n\n")
