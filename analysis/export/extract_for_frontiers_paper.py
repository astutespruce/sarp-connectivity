import os
from pathlib import Path

import geopandas as gp
import pandas as pd
from openpyxl.styles import Alignment, NamedStyle
from openpyxl.utils import get_column_letter

from analysis.constants import (
    SARP_STATE_NAMES,
    ONSTREAM_MANUALREVIEW,
)
from api.constants import DOMAINS, unpack_domains

primary_col_style = NamedStyle(
    name="PrimaryColumnStyle", alignment=Alignment(horizontal="left", wrap_text=True)
)

version = "12_17_2021"

data_dir = Path("data/barriers/master")
out_dir = Path("/tmp/sarp")

if not out_dir.exists():
    os.makedirs(out_dir)

had_manual_review = [4, 5, 8, 10, 11, 13, 14, 15]


barrier_type = "dams"

df = gp.read_feather(data_dir / f"{barrier_type}.feather").set_index("id")
api_df = pd.read_feather(f"data/api/{barrier_type}.feather").set_index("id")

df = df.join(api_df[["HasNetwork", "Ranked"]]).rename(
    columns={
        "unranked": "invasive",  # NOTE: IMPORTANT: this field now includes diversions
        "HasNetwork": "has_networkresults",
        "Ranked": "ranked",
        "ProtectedLand": "on_protectedland",
        "loop": "on_loop",
    }
)

# Exclude dams that have already been removed; these are not included in Frontiers stats
df = df.loc[~((df.Recon == 7) | (df.Feasibility == 8) | (df.ManualReview == 8))].copy()


df["in_NID"] = df.NIDID != ""
df["analyzed"] = df.snapped & ~(df.duplicate | df.dropped | df.excluded)
df["is_reconned"] = (df.Recon > 0).astype("uint8")
df["is_error"] = df.dropped | df.duplicate
df["is_manualreviewed"] = df.ManualReview.isin(had_manual_review).astype("uint8")
df["is_manualreviewed_onnetwork"] = df.ManualReview.isin(ONSTREAM_MANUALREVIEW)
df["is_manualreviewed_offnetwork"] = df.ManualReview == 5
df["is_manualreviewed_error"] = df.ManualReview.isin([6, 11, 14])
df["is_estimated_manualreviewed"] = df.is_estimated & df.is_manualreviewed
df["is_estimated_reconned"] = df.is_estimated & df.is_reconned
df["is_estimated_reconned_or_manualreviewed"] = df.is_estimated & (
    df.is_reconned | df.is_manualreviewed
)
df["is_estimated_error"] = df.is_estimated & df.is_error


status_fields = [
    "in_NID",
    "dropped",
    "excluded",
    "duplicate",
    "snapped",
    "analyzed",
    "has_networkresults",
    "ranked",
    "is_manualreviewed",
    "is_manualreviewed_onnetwork",
    "is_manualreviewed_offnetwork",
    "is_manualreviewed_error",
    "is_reconned",
    "is_error",
    "invasive",
    "is_estimated",
    "is_estimated_manualreviewed",
    "is_estimated_reconned",
    "is_estimated_reconned_or_manualreviewed",
    "is_estimated_error",
    "on_protectedland",
    "on_loop",
]

summary_fields = [
    "Recon",
    "Feasibility",
    "ManualReview",
    "log",
    "snap_log",
]


for col in status_fields + ["OwnerType"]:
    df[col] = df[col].fillna(0).astype("uint8")

# limit to SARP states, exclude estimated 2021
df = df.loc[
    df.State.isin(SARP_STATE_NAMES) & (~df.Source.isin(["Estimated Dams OCT 2021"]))
].copy()

# write out raw CSV for archive for Frontiers paper, with domain values converted text
archive = pd.DataFrame(
    df[
        [
            "lat",
            "lon",
            "SARPID",
            "Recon",
            "Feasibility",
            "ManualReview",
            "OwnerType",
            "in_NID",
            "dropped",
            "excluded",
            "analyzed",
            "has_networkresults",
            "ranked",
            "is_manualreviewed",
            "is_manualreviewed_onnetwork",
            "is_manualreviewed_offnetwork",
            "is_manualreviewed_error",
            "is_reconned",
            "is_error",
            "invasive",
            "is_estimated",
            "is_estimated_manualreviewed",
            "is_estimated_reconned",
            "is_estimated_reconned_or_manualreviewed",
            "is_estimated_error",
            "on_protectedland",
            "on_loop",
        ]
    ].reset_index()
)
# archive["ProtectedLand"] = archive.ProtectedLand.astype("bool")
archive = unpack_domains(archive)
archive.to_csv(out_dir / f"{barrier_type}_{version}_frontiers_archive.csv", index=False)


# status_fields += ["ProtectedLand"]
summary_fields += ["OwnerType"]

states = sorted(x for x in df.State.unique() if x)
huc2 = sorted(x for x in df.HUC2.unique() if x)
df = df.reset_index()


with pd.ExcelWriter(
    out_dir / f"{barrier_type}_{version}_summary_for_frontiers_manuscript.xlsx"
) as xlsx:

    ### Totals
    data = {
        "total": len(df),
        "in_NID": df.in_NID.sum(),
        "dropped": df.dropped.sum(),
        "excluded": df.excluded.sum(),
        "duplicate": df.duplicate.sum(),
        "snapped": df.snapped.sum(),
        "analyzed": df.analyzed.sum(),
        "has_networkresults": df.has_networkresults.sum(),
        "ranked": df.ranked.sum(),
        "is_manualreviewed": df.is_manualreviewed.sum(),
        "is_manualreviewed_onnetwork": df.is_manualreviewed_onnetwork.sum(),
        "is_manualreviewed_offnetwork": df.is_manualreviewed_offnetwork.sum(),
        "is_manualreviewed_error": df.is_manualreviewed_error.sum(),
        "is_reconned": df.is_reconned.sum(),
        "is_error (dropped or duplicate)": df.is_error.sum(),
        "invasive": df.invasive.sum(),
        "is_estimated": df.is_estimated.sum(),
        "is_estimated_manualreviewed": df.is_estimated_manualreviewed.sum(),
        "is_estimated_reconned": df.is_estimated_reconned.sum(),
        "is_estimated_reconned_or_manualreviewed": df.is_estimated_reconned_or_manualreviewed.sum(),
        "is_estimated_error": df.is_estimated_error.sum(),
        "on_protectedland": df.on_protectedland.sum(),
        "on_loop": df.on_loop.sum(),
    }

    total = pd.DataFrame([data]).melt()
    total.columns = ["", "count"]
    sheet_name = f"All {barrier_type}"
    total.to_excel(xlsx, sheet_name=sheet_name, index=False)
    ws = xlsx.sheets[sheet_name]
    ws.column_dimensions["A"].width = 32

    ### Totals by state
    agg = {
        "id": "count",
        "in_NID": "sum",
        "dropped": "sum",
        "excluded": "sum",
        "duplicate": "sum",
        "snapped": "sum",
        "analyzed": "sum",
        "has_networkresults": "sum",
        "ranked": "sum",
        "is_manualreviewed": "sum",
        "is_manualreviewed_onnetwork": "sum",
        "is_manualreviewed_offnetwork": "sum",
        "is_manualreviewed_error": "sum",
        "is_reconned": "sum",
        "is_error": "sum",
        "is_estimated": "sum",
        "is_estimated_manualreviewed": "sum",
        "is_estimated_reconned": "sum",
        "is_estimated_reconned_or_manualreviewed": "sum",
        "is_estimated_error": "sum",
        "on_loop": "sum",
    }

    stats = df.groupby("State").agg(agg)
    sheet_name = f"{barrier_type} by state"
    stats.to_excel(xlsx, sheet_name=sheet_name)
    ws = xlsx.sheets[sheet_name]
    ws.column_dimensions["A"].width = 24
    for idx in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 16

    for row in list(ws.rows)[1:]:
        row[0].style = primary_col_style

    ### Totals by HUC2
    stats = df.groupby("HUC2").agg(agg)
    sheet_name = f"{barrier_type} by HUC2"
    stats.to_excel(xlsx, sheet_name=sheet_name)
    ws = xlsx.sheets[sheet_name]
    ws.column_dimensions["A"].width = 6
    for idx in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 16

    for row in list(ws.rows)[1:]:
        row[0].style = primary_col_style

    for col in summary_fields:
        stats = df.groupby(col).agg(agg)

        if col in DOMAINS:
            stats.index = (
                stats.index.astype(str)
                + ": "
                + stats.index.map(DOMAINS[col]).fillna("")
            )

        sheet_name = f"{col} {barrier_type}"
        stats.to_excel(xlsx, sheet_name=sheet_name)
        ws = xlsx.sheets[sheet_name]
        ws.column_dimensions["A"].width = 60
        for idx in range(2, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 16

        for row in list(ws.rows)[1:]:
            row[0].style = primary_col_style

        ### Crosstab by state
        values = df[col].copy()
        if col in DOMAINS:
            values = df[col].astype(str) + ": " + df[col].map(DOMAINS[col]).fillna("")

        crosstab = pd.crosstab(
            values,
            df.State,
        )

        sheet_name = f"{col} {barrier_type} state crosstab (total)"
        crosstab.to_excel(xlsx, sheet_name=sheet_name)
        ws = xlsx.sheets[sheet_name]
        ws.column_dimensions["A"].width = 60
        for idx in range(2, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 16

        for row in list(ws.rows)[1:]:
            row[0].style = primary_col_style

        ### Values by state
        stats = df.groupby(["State", col]).agg(agg).reset_index()
        if col in DOMAINS:
            stats[col] = (
                stats[col].astype(str) + ": " + stats[col].map(DOMAINS[col]).fillna("")
            )

        sheet_name = f"{col} {barrier_type} by state"
        stats.to_excel(xlsx, sheet_name=sheet_name, index=False)
        ws = xlsx.sheets[sheet_name]
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 60
        for idx in range(3, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 16

        for row in list(ws.rows)[1:]:
            row[0].style = primary_col_style
            row[1].style = primary_col_style

        ### Crosstab by HUC2
        values = df[col].copy()
        if col in DOMAINS:
            values = df[col].astype(str) + ": " + df[col].map(DOMAINS[col]).fillna("")

        crosstab = pd.crosstab(values, df.HUC2)

        sheet_name = f"{col} {barrier_type} HUC2 crosstab (total)"
        crosstab.to_excel(xlsx, sheet_name=sheet_name)
        ws = xlsx.sheets[sheet_name]
        ws.column_dimensions["A"].width = 60
        for idx in range(2, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 16

        for row in list(ws.rows)[1:]:
            row[0].style = primary_col_style

        ### Values by HUC2
        stats = df.groupby(["HUC2", col]).agg(agg).reset_index()
        if col in DOMAINS:
            stats[col] = (
                stats[col].astype(str) + ": " + stats[col].map(DOMAINS[col]).fillna("")
            )

        sheet_name = f"{col} {barrier_type} by HUC2"
        stats.to_excel(xlsx, sheet_name=sheet_name, index=False)
        ws = xlsx.sheets[sheet_name]
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 60
        for idx in range(3, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 16

        for row in list(ws.rows)[1:]:
            row[0].style = primary_col_style
            row[1].style = primary_col_style
